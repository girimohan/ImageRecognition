from keras.applications import VGG16
from keras.applications import imagenet_utils
from tensorflow.keras.utils import img_to_array, load_img
import numpy as np
import cv2
import os
import openpyxl

#loading the model
pretrained_model = VGG16(weights="imagenet")

# create a new Excel workbook
wb = openpyxl.Workbook()
ws = wb.active

# add column headers
ws.cell(row=1, column=1, value='Image Name')
ws.cell(row=1, column=2, value='Prediction')
ws.cell(row=1, column=3, value='Category')
ws.cell(row=1, column=4, value='Accuracy')

# loading the images from the directory
img_dir = 'images/pictures'
for img_name in os.listdir(img_dir):
    img_path = os.path.join(img_dir, img_name)
    img = load_img(img_path, target_size=(224, 224))

    # convert the image into a 4 dimensional Tensor
    # convert from (height, width, channels), (batchsize, height, width, channels)
    img_array = np.expand_dims(img_to_array(img), axis=0)

    # preprocess the input image array
    img_array = imagenet_utils.preprocess_input(img_array)

    # predict using predict() method
    prediction = pretrained_model.predict(img_array)

    # decode the predictions and get the top predicted category and its accuracy
    #decoded_predictions = imagenet_utils.decode_predictions(prediction, top=1000)
    decoded_predictions = imagenet_utils.decode_predictions(prediction)

    top_category = decoded_predictions[0][0][1]
    accuracy = decoded_predictions[0][0][2]*100

    # classify the top predicted category into the three categories
    if top_category == 'space_shuttle'or top_category == 'airliner':
        category = 'aeroplane'
    elif top_category == 'minivan' or top_category == 'jeep'or top_category == 'tow_truck' or top_category == 'limousine':
        category = 'car'
    elif top_category == 'moped' or top_category == 'motor_scooter' or top_category == 'mountain_bike':
        category = 'motorbike'
    else:
        category = 'other'

    # add the prediction result to the Excel worksheet
    ws.append([img_name, top_category, category, accuracy])

    # display image and the prediction text over it
    # disp_img = cv2.imread(img_path)
    # # display prediction text over the image
    # cv2.putText(disp_img, top_category, (20, 20), cv2.FONT_HERSHEY_TRIPLEX , 0.8, (255, 0, 0))
    # # show the image
    # cv2.imshow("Prediction", disp_img)
    # cv2.waitKey()

# save the Excel workbook
wb.save('prediction_results.xlsx')
